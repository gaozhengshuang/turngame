#!/usr/bin/env python
#coding=utf-8
import os
import sys
import string
import openpyxl
import openpyxl.worksheet.dimensions

#print openpyxl.__file__
class GenProtoIndex:

    ## init
    def __init__(self, filepath, outfile):
        self.filepath = filepath
        self.outfile = outfile
        self.listfile = []
        self.listfp = []
        print filepath, outfile

    ## parse .proto file
    def ParsePath(self):
        flist = os.listdir(self.filepath)
        for f in flist:
            fullpath = os.path.join(self.filepath, f)
            if os.path.isfile(fullpath) == False:
                continue
            if os.path.splitext(fullpath)[1] != ".proto":
                continue
            self.listfile.append(fullpath)

    ##
    def ParseFiles(self):
        for f in self.listfile:
            print "==>>开始分析文件 %s\n" % f,
            fd = open(f)
            line = fd.readline()
            fp = FileProtobuf()
            while line:
                self.ParseLine(line, fp)
                line = fd.readline()
            self.listfp.append(fp)

    ##
    def ParseLine(self, line, fp):
        elements = string.split(line," ")
        if len(elements) == 0:
            return

        if elements[0] == "package":
            pgkname = string.split(elements[1], ";")[0]
            fp.pkg = pgkname

        if elements[0] == "message":
            msgname = elements[1]
            fp.messages.append(msgname)

    ##
    def WriteExcelFile(self):
        wb = openpyxl.Workbook()

        ws = wb.worksheets[0]
        ws.title = "proto_index"

        ws.cell(row=1, column=1).value = "Id"
        ws.cell(row=1, column=2).value = "Name"
        ws.cell(row=2, column=1).value = "int32"
        ws.cell(row=2, column=2).value = "string"
        ws.cell(row=3, column=1).value = '''RepeatCheck:true MakeIndex:true json:"id"'''
        ws.cell(row=3, column=2).value = '''RepeatCheck:true MakeIndex:true json: "name"'''
        ws.cell(row=4, column=1).value = "本文件自动生成"

        databegin, uuid = 5, 1
        for fp in self.listfp:
            for msg in fp.messages:
                c1 = ws.cell(databegin+uuid-1, 1)
                c1.value = uuid
                
                fullmsg = fp.pkg + "." + msg
                c2 = ws.cell(databegin+uuid-1, 2)
                c2.value = fullmsg
                uuid += 1
    
        #
        #style_font = openpyxl.styles.Font(name='Verdana', size=12, color='FF123456')

        self.WriteExcelSecondSheet(wb)
        wb.save(self.outfile)

        ### 工作薄数据
        #print wb.worksheets
        #print wb.style_names
        #print wb.active

        ### 表单sheet数据
        # 按列遍历
        #for col in ws.iter_cols():
        #    for cell in col:
        #        print cell.value

        # 按行遍历
        #for row in ws.iter_rows():
        #    for cell in row:
        #        print cell.value

        # 尺寸
        #print ws.dimensions


        print"==>>生成文件 %s\n" % self.outfile

    def WriteExcelSecondSheet(self, wb):
        ws = wb.create_sheet() 
        ws.title = "@Types"
        cell = ws.cell(1,1)
        cell.value = '''TableName: "ProtoId" Package: "table" CSClassHeader: "[System.Serializable]"'''

        ws.cell(2,1).value = "ObjectType"
        ws.cell(2,2).value = "FieldName"
        ws.cell(2,3).value = "FieldType"
        ws.cell(2,4).value = "Value"
        ws.cell(2,5).value = "Alias"
        ws.cell(2,6).value = "Default"
        ws.cell(2,7).value = "Meta"
        ws.cell(2,8).value = "Comment"

        ws.cell(3,1).value = "对象类型"
        ws.cell(3,2).value = "字段名"
        ws.cell(3,3).value = "字段类型"
        ws.cell(3,4).value = "枚举值"
        ws.cell(3,5).value = "别名"
        ws.cell(3,6).value = "默认值"
        ws.cell(3,7).value = "特性"
        ws.cell(3,8).value = "注释"


##
if len(sys.argv) != 3:
    print "please input .proto file path and output excel file!"
    exit()

class FileProtobuf:
    def __init__(self):
        self.pkg = ''     # 名称
        self.messages = []     # 列表

##
filepath = string.split(sys.argv[1],"=")
outfile = string.split(sys.argv[2],"=")

impl = GenProtoIndex(filepath[1], outfile[1])
impl.ParsePath()
impl.ParseFiles()
impl.WriteExcelFile()


